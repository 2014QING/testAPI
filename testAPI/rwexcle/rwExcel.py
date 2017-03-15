# -*- coding: utf-8 -*-
import xlrd
from openpyxl import load_workbook
import time
from datetime import date, datetime


class RwExcel:
    def __int__(self):
        pass

    def open_excel(self, file):
        try:
            data = xlrd.open_workbook(file)
            return data
        except Exception as e:
            print(str(e))

    # 根据名称获取Excel表格中的数据   参数:file：Excel文件路径 ，colnameindex：表头列名所在行的索引，by_name：Sheet1名称
    def read_excel_table_byindex(self, file, colnameindex=0, by_index=0):
        book = self.open_excel(file)
        table = book.sheet_by_index(by_index)
        nrows = table.nrows  # 行数
        ncols = table.ncols  # 列数
        #python读取excel中单元格的内容返回的有5种类型，即ctype:ctype :  0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
        colnames = table.row_values(colnameindex)  # 第0行数据
        list_data = []
        for rownum in range(1, nrows):
            row = table.row_values(rownum)  # 每行数据
            if row:
                app = {}
                for i in range(len(colnames)):
                    # table.cell(rownum,i).ctype==3，说明是时间格式的值
                    if (table.cell(rownum, i).ctype == 3):
                        date_value = xlrd.xldate_as_tuple(table.cell(rownum, i).value, book.datemode)
                        app[colnames[i]] = date(*date_value[:3]).strftime('%Y/%m/%d')
                    else:
                        app[colnames[i]] = row[i]  # 第0行数据表头key:列值value，json格式
                list_data.append(app)
        return list_data

    def main(self):
        tables = self.read_excel_table_byindex()
        for row in tables:
            print(row)
            print(row['出发地'])

    # 修改方法不成功，报错，可能xlutils不支持python3.5
    # def edit_excel_table_byindex(self,file,rowindex,colindex,newvalue,by_index=0):
    #     book=self.open_excel(file)
    #     #book=xlrd.open_workbook(file)
    #     table = book.sheet_by_index(by_index)
    #     tmpBook=copy(book)
    #     ws=tmpBook.get_sheet(0)
    #     ws.write(rowindex,colindex,newvalue)
    #     tmpBook.save('D:\\MyCodes\\shopping\\AVtestdatasResult.xlsx')

    #def write_excel_table_byindex(self, rowindex, colindex, listnewvalue,file,filersname, by_index=0):
        # 加载一个已经存在的excel
    #    wb = load_workbook(file)
     #   sheet = wb.active
        # row和column都是从1开始
    #    for value in listnewvalue:
    #       sheet.cell(row=rowindex, column=colindex).value = value
    #        rowindex = rowindex + 1
    #    time.sleep(5)
     #   wb.save('/Users/zhaoqing/softwares/codes/testAPI/scripts/params/'+filersname+ time.strftime('%Y-%m-%d-%H-%M-%S',time.localtime(time.time())) + '.xlsx')





        # if __name__=="__main__":
        #     test=RwEXcel()
        #     test.main()
        #     test.write_excel_table_byindex(rowindex=2,colindex=22,newvalue='测试结果')

    def write_excel_table_byindex(self,title, listnewvalue,
                                  file, filersname, by_index=0):
        # 加载一个已经存在的excel
        wb = load_workbook(file)
        sheet = wb.active
        row = 1
        col=self.return_col(file,title)
        # row和column都是从1开始
        rowindex=2
        for value in listnewvalue:
            sheet.cell(row=rowindex, column=col).value = value
            rowindex = rowindex + 1
        time.sleep(5)
        wb.save('/Users/zhaoqing/softwares/codes/testAPI/scripts/params/' + filersname + time.strftime('%Y-%m-%d-%H-%M-%S',
                                                                                                   time.localtime(time.time())) + '.xlsx')

    def write_result_table_byindex(self, title, listnewvalue,
                                      file, filersname, by_index=0):
        # 加载一个已经存在的excel
        wb = load_workbook(file)
        sheet = wb.active
        row = 1
        col = self.return_col(file, title)
        # row和column都是从1开始
        rowindex = 2
        for value in listnewvalue:
            sheet.cell(row=rowindex, column=col).value = value
            rowindex = rowindex + 1
            time.sleep(5)
        wb.save('/Users/zhaoqing/softwares/codes/testAPI/scripts/params/' + filersname + time.strftime(
                '%Y-%m-%d-%H-%M-%S',time.localtime(time.time())) + '.xlsx')


    # 返回file所在的列（表头一行）
    def return_col(self,file,title,colnameindex=0,by_index=0):
        book = self.open_excel(file)
        table = book.sheet_by_index(by_index)
        col=0
        colnames = table.row_values(colnameindex)  # 第0行数据,表头
        for i in range(len(colnames)):
            try:
                if(colnames[i]==title ):
                    col=i+1
                    return col
            except Exception as e:
                print("error:",e)
                raise
        if(col==0):
            print("Don't have this title")


if __name__=="__main__":
    test=RwExcel()
    #test.main()
             #test.write_excel_table_byindex(rowindex=2,colindex=22,newvalue='测试结果')
    file = "/Users/zhaoqing/softwares/codes/testAPI/scripts/params/testaddOW.xlsx"
    col=test.return_col(file,'excepted_rs')
    print("*****",col)