# -*- coding:utf-8 -*-

import os
import shutil
import tempfile
from lxml import etree
from shopping.flightSegment import flightSegment
from rwexcle.rwExcel import RwExcel
from openpyxl import load_workbook
import time
import operator


class CompShoppingDatas:
    def __int__(self):
        pass
        # self.rwexcel= RwExcel()

    def newfile(self):
        temp_file = tempfile.mktemp()
        ffopen = open(temp_file, 'w')
        old_file = '/Users/zhaoqing/Downloads/shopping_rs.xml'
        fopen = open(old_file, 'r')
        for line in fopen:
            if line[0] == '-':
                line = line.replace('- <', '<')
            ffopen.write(line)
        fopen.close()
        ffopen.close()
        if os.path.exists(old_file):
            os.remove(old_file)
        shutil.copy(temp_file, old_file)
        try:
            os.remove(temp_file)
        except OSError:
            print(OSError)
        print('OK')

    def write_result_table_byindex(self, collist, valuelist, file, filersname, by_index=0):
        # 加载一个已经存在的excel
        wb = load_workbook(file)
        sheet = wb.active
        # row和column都是从1开始,row第一行为标题，从第二行开始写入数据
        rowindex = 2
        for rs in valuelist:
            # print('eeeeeeeeetttttt',rs)
            if rs['code'] == 'OSError':
                rowindex = rowindex + 1
                continue
            if rs['code'] == 'OK':
                rsA = str(rs['result_A'])
                coA = str(rs['count_A'])
                rsB = str(rs['result_B'])
                coB = str(rs['count_B'])
                rsC = str(rs['result_C'])
                coC = str(rs['count_C'])
                rsD = str(rs['result_D'])
                coD = str(rs['count_D'])
                coNew = str(rs['count_new'])
                coOld = str(rs['count_old'])
                for col in collist:
                    if col == 4:
                        sheet.cell(row=rowindex, column=col).value = rsA
                    elif col == 5:
                        sheet.cell(row=rowindex, column=col).value = coA
                    elif col == 6:
                        sheet.cell(row=rowindex, column=col).value = rsB
                    elif col == 7:
                        sheet.cell(row=rowindex, column=col).value = coB
                    elif col == 8:
                        sheet.cell(row=rowindex, column=col).value = rsC
                    elif col == 9:
                        sheet.cell(row=rowindex, column=col).value = coC
                    elif col == 10:
                        sheet.cell(row=rowindex, column=col).value = rsD
                    elif col == 11:
                        sheet.cell(row=rowindex, column=col).value = coD
                    elif col == 12:
                        sheet.cell(row=rowindex, column=col).value = coNew
                    elif col == 13:
                        sheet.cell(row=rowindex, column=col).value = coOld
                rowindex = rowindex + 1

        time.sleep(5)
        wb.save(
            '/Users/zhaoqing/softwares/codes/testAPI/scripts/params/shopping/Aresults/' + filersname + time.strftime(
                '%Y-%m-%d-%H-%M-%S', time.localtime(time.time())) + '.xlsx')

    def readxml(self, dep, arr, xmlfile):
        flightDict = {}
        try:
            tree =etree.parse(xmlfile)
            root = tree.getroot()

            rs = []

            # 将flightID,flightNumber存入flightDict字典中
            for child1 in root:
                if child1.tag == 'searchResult':
                    for child12 in child1:
                        if child12.tag == 'flightList':
                            for child13 in child12:
                                if child13.tag == 'flight':
                                    # flightID,flightNumber
                                    flightDict[child13[0].text] = child13[2].text
            for child1 in root:
                if child1.tag == 'searchResult':
                    for child12 in child1:
                        if child12.tag == 'flightProductGroupList':
                            for child14 in child12:
                                flightIDList = []
                                flightIDDict = {}
                                bookingClassList = []
                                if child14.tag == 'flightProductGroup':
                                    # 将flightSegmentList和priceList放在一个对象中，将所有对象放在list中
                                    for child15 in child14:
                                        fs = flightSegment()
                                        if child15.tag == 'flightSegmentList':
                                            # 循环filghtSegment中的数据
                                            # child15为flightSegmentList的总集合
                                            for child16 in child15:
                                                # 将flightID存入list中
                                                # flightIDDict[child16[3].text] = flightDict[child16[3].text]
                                                # if flightIDDict not in flightIDList:
                                                #     flightIDList.append(flightIDDict)
                                                flightID = flightDict[child16[3].text]
                                                if flightID not in flightIDList:
                                                    flightIDList.append(flightID)
                                                fs.setdepCity(dep)
                                                fs.setarrCity(arr)
                                        fs.setflightID(flightIDList)
                                        # fs.setflightID(flightIDDict)
                                        if child15.tag == 'priceList':
                                            fareList = []
                                            for child17 in child15:
                                                for child18 in child17:
                                                    fare = {}
                                                    bookingClassInfo = {}
                                                    if child18.tag == 'priceInfo':
                                                        #for child20 in child18:
                                                            #fare=etree.Element('fare')
                                                            #fare=child20.get("fare")
                                                            #print('kkkkkkkkk',fare)
                                                        # child18[0].text为travelerCategoryCode的值，child18[2].text为fare的值
                                                        if child18[2].text=='CNY':
                                                            fare['travelerCategoryCode'] = child18[0].text
                                                            fare['fare'] = child18[3].text
                                                        else:
                                                            fare['travelerCategoryCode'] = child18[0].text
                                                            fare['fare'] = child18[2].text
                                                        fareList.append(fare)
                                                        fs.setfare(fareList)
                                                    if child18.tag == 'bookingClassInfoList':
                                                        for child19 in child18:
                                                            # child19[0].text为bookingClass的值，child19[1].text为cabinClass的值
                                                            #bookingClassInfo['bookingClass'] = child19[0].text
                                                            bookingClassInfo['cabinClass'] = child19[1].text
                                                            bookingClassList.append(bookingClassInfo)
                                                            fs.setcabinClass(bookingClassList)
                                rs.append(fs)

            return rs
        #往返的时候，有些OD只有去，没有回，读取会报错，提示没有该文档，故添加报错

        except OSError:
            errors = 1
            return errors

    def compfare(self, file_od, file1, file2, filename, type):

        self.rwexcel = RwExcel()
        odDatas = self.rwexcel.read_excel_table_byindex(file_od)
        rsList = []

        for od in odDatas:
            # 相同的航班组合fare值完全一致 （A）
            # 相同的航班组合fare值不一致(B)
            # 灰度之前的航班组合在灰度之后未找到匹配的航班组合(C)
            # 灰度之后的航班组合在灰度之前未找到匹配的航班组合(D)
            result_A = [[od['DEP'], od['ARR']]]
            result_B = [[od['DEP'], od['ARR']]]
            result_C = [[od['DEP'], od['ARR']]]
            result_D = [[od['DEP'], od['ARR']]]
            flight_new = []
            flight_old = []
            rsDict = {}
            count_A = 0
            count_B = 0
            if type == 'OW':
                file_old = file1 + od['DEP'] + '-' + od['ARR'] + '.xml'
                file_new = file2 + od['DEP'] + '-' + od['ARR'] + '.xml'
            elif type == 'RT':
                file_old = file1 + od['DEP'] + '-' + od['ARR'] + '-return.xml'
                file_new = file2 + od['DEP'] + '-' + od['ARR'] + '-return.xml'
            rs_old = self.readxml(od['DEP'], od['ARR'], file_old)
            rs_new = self.readxml(od['DEP'], od['ARR'], file_new)
            if rs_new == 1 or rs_old == 1:
                #rs_new和rs_old 无正常结果时，code='OSError'，跳出循环，继续下一次循环
                rsList.append({'code': 'OSError'})
                continue
            rsDict['code'] = 'OK'
            count_new = len(rs_new)
            count_old = len(rs_old)
            for item_new in rs_new:
                # flight_new为灰度后每个OD的航班组合
                flight_new.append(item_new.getflightID())
                for item_old in rs_old:
                    # 去掉重复flightNumber
                    if item_old.getflightID() not in flight_old:
                        #flight_old为灰度前每个OD的航班组合，因为在for item_new in rs_new的循环下，所以需要去重
                        flight_old.append(item_old.getflightID())
                    # 需要将item_new.getflightID()，item_new.getcabinClass()，item_new.getfare()排序，让新老结果的顺序一致，比较的时候，不会因为顺序不一致(实际元素是相等的)，导致比较结果有问题
                    # tem_new.getflightID()是list类型，item_new.getcabinClass()，item_new.getfare()是元素为dict的list数据
                    if sorted(item_new.getflightID()) == sorted(item_old.getflightID()):
                        if sorted(item_new.getcabinClass(), key=operator.itemgetter('cabinClass')) == sorted(item_old.getcabinClass(), key=operator.itemgetter('cabinClass')):
                            if sorted(item_new.getfare(), key=operator.itemgetter('fare')) == sorted(item_old.getfare(), key=operator.itemgetter('fare')):
                                count_A = count_A + 1
                                #在result_A中去掉重复flightID
                                if item_new.getflightID() not in result_A:
                                    result_A.append(item_new.getflightID() + item_new.getfare())
                            else:
                                count_B = count_B + 1
                                result_B.append(item_new.getflightID() +item_new.getfare() + item_old.getfare())
                        #可能存在航班组合一样，但是价格及对应的舱等个数不一致的情况，放到B结果中
                        else:
                            count_B = count_B + 1
                            result_B.append(item_new.getflightID() + item_new.getcabinClass() + item_new.getfare()  +item_old.getfare() + item_old.getcabinClass())
            rsDict['result_A'] = result_A
            rsDict['result_B'] = result_B
            result_C.append([item for item in flight_old if item not in flight_new])
            result_D.append([item for item in flight_new if item not in flight_old])
            count_C = len(result_C[1])
            count_D = len(result_D[1])
            rsDict['result_C'] = result_C
            rsDict['result_D'] = result_D
            rsDict['count_A'] = count_A
            rsDict['count_B'] = count_B
            rsDict['count_C'] = count_C
            rsDict['count_D'] = count_D
            rsDict['count_new'] = count_new
            rsDict['count_old'] = count_old
            rsList.append(rsDict)
        self.write_result_table_byindex([4, 5, 6, 7, 8, 9, 10, 11, 12, 13], rsList, file_od, filename)




if __name__ == '__main__':
    rx = CompShoppingDatas()
    # rx.newfile()
    # rx.readxml()
    file_od = '/Users/zhaoqing/softwares/codes/testAPI/scripts/params/shopping/od.xlsx'
    file_old = '/Users/zhaoqing/softwares/codes/testAPI/scripts/params/shopping/shoppingold20170314/'
    file_new = '/Users/zhaoqing/softwares/codes/testAPI/scripts/params/shopping/shoppingnew20170314/'
    filename1 = 'shopingOW'
    filename2 = 'shopingRT'
    rx.compfare(file_od, file_old, file_new, filename1, 'OW')
    rx.compfare(file_od, file_old, file_new, filename2, 'RT')
